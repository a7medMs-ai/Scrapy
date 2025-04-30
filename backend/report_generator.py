# backend/report_generator.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import os


def generate_excel_report(data_list, report_path, lang_code="general"):
    # تحويل البيانات إلى DataFrame
    df = pd.DataFrame(data_list)

    # إنشاء ملف Excel جديد أو فتحه
    if not os.path.exists(report_path):
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            _write_sheet(writer, df, lang_code)
    else:
        with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            _write_sheet(writer, df, lang_code)


def _write_sheet(writer, df, sheet_name):
    # أضف المجموع في آخر الصفوف
    totals = {
        "Page Counter": "TOTAL",
        "Word Count": df["Word Count"].sum(),
        "Segments": df["Segments"].sum()
    }
    df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # تنسيق إضافي للـ TOTAL row
    writer.book = writer.book if writer.book else load_workbook(writer.path)
    sheet = writer.book[sheet_name]
    last_row = sheet.max_row

    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=last_row, column=col)
        if cell.value == "TOTAL":
            cell.font = Font(bold=True, color="FF0000")  # أحمر عريض
